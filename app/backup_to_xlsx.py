#!/usr/bin/env python3
"""
backup_to_xlsx.py
------------------
Converts a "Cashew"-style finance-app backup JSON export into an Excel
workbook: one Transactions sheet plus one lookup sheet per referenced
entity (Accounts, Categories, Tags, Places, People, Users, Recurrings,
Budgets, Loans). Every entity reference in the Transactions sheet is a
live INDEX/MATCH formula against the relevant lookup sheet (by UUID),
so there's no duplicated/hardcoded text and the file stays correct if
you fix a name in the lookup sheet.

USAGE
    python3 backup_to_xlsx.py <input_backup.json> [output.xlsx]

If output.xlsx is omitted, it defaults to <input_backup>.xlsx next to
the input file.

Only sheets/columns for entities that actually appear with data in the
given backup are created, so this script adapts automatically to
different exports (e.g. one where Budgets or Loans are actually used).
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------
# styling helpers
# ---------------------------------------------------------------------
def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_font_and_border(ws, nrows, ncols, start_row=2):
    for r in range(start_row, start_row + nrows):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.border = BORDER


# ---------------------------------------------------------------------
# generic lookup-sheet builder
# ---------------------------------------------------------------------
def build_lookup_sheet(wb, sheet_name, items, columns):
    """
    items: list of dicts (raw entities from the backup JSON)
    columns: list of (header, extractor_fn) tuples. First column MUST
             be the UUID (used as the MATCH key from the main sheet).
    Returns (sheet_name, n_rows_incl_header) or None if items is empty.
    """
    if not items:
        return None
    ws = wb.create_sheet(sheet_name)
    headers = [h for h, _ in columns]
    ws.append(headers)
    style_header(ws, len(headers))
    for item in items:
        ws.append([fn(item) for _, fn in columns])
    apply_font_and_border(ws, len(items), len(headers))
    widths = [38] + [24] * (len(headers) - 1)
    autosize(ws, widths)
    ws.freeze_panes = "A2"
    return sheet_name, len(items) + 1  # +1 for header row


# ---------------------------------------------------------------------
# main conversion
# ---------------------------------------------------------------------
def convert(input_path: Path, output_path: Path):
    with open(input_path) as f:
        data = json.load(f)

    transactions = data.get("transactions") or []
    if not transactions:
        raise ValueError("No transactions found in backup JSON.")

    wb = Workbook()
    wb.remove(wb.active)

    acc_type_map = {0: "Card", 1: "Cash", 2: "Bank Account"}
    cat_type_map = {0: "Expense", 1: "Income"}
    txn_type_map = {0: "Expense", 1: "Income", 2: "Transfer"}

    lookups = {}  # key -> (sheet_name, n_rows)

    lookups["account"] = build_lookup_sheet(
        wb, "Accounts", data.get("accounts") or [],
        [
            ("UUID", lambda a: a.get("uuid")),
            ("Account Name", lambda a: a.get("name")),
            ("Bank Name", lambda a: a.get("bankName") or ""),
            ("Account Type", lambda a: acc_type_map.get(a.get("type"), a.get("type"))),
            ("Currency Code", lambda a: a.get("currencyCode") or ""),
            ("Balance", lambda a: a.get("amount")),
            ("Is Default", lambda a: "Yes" if a.get("isDefault") else "No"),
            ("Is Excluded", lambda a: "Yes" if a.get("isExcluded") else "No"),
        ],
    )
    lookups["category"] = build_lookup_sheet(
        wb, "Categories", data.get("categories") or [],
        [
            ("UUID", lambda c: c.get("uuid")),
            ("Category Name", lambda c: c.get("name")),
            ("Category Kind", lambda c: cat_type_map.get(c.get("type"), c.get("type"))),
            ("Description", lambda c: c.get("description") or ""),
        ],
    )
    lookups["tag"] = build_lookup_sheet(
        wb, "Tags", data.get("labels") or [],
        [
            ("UUID", lambda t: t.get("uuid")),
            ("Tag Name", lambda t: t.get("name")),
        ],
    )
    lookups["place"] = build_lookup_sheet(
        wb, "Places", data.get("places") or [],
        [
            ("UUID", lambda p: p.get("uuid")),
            ("Place Name", lambda p: p.get("name")),
            ("Description", lambda p: p.get("description") or ""),
        ],
    )
    lookups["people"] = build_lookup_sheet(
        wb, "People", data.get("peoples") or [],
        [
            ("UUID", lambda p: p.get("uuid")),
            ("Person Name", lambda p: p.get("name")),
            ("Description", lambda p: p.get("description") or ""),
        ],
    )
    lookups["user"] = build_lookup_sheet(
        wb, "Users", data.get("users") or [],
        [
            ("UUID", lambda u: u.get("uuid")),
            ("User Name", lambda u: u.get("name")),
            ("Currency", lambda u: u.get("currency") or ""),
        ],
    )
    lookups["recurring"] = build_lookup_sheet(
        wb, "Recurrings", data.get("recurrings") or [],
        [
            ("UUID", lambda r: r.get("uuid")),
            ("Recurring Name", lambda r: r.get("name")),
            ("Period", lambda r: r.get("period")),
        ],
    )
    lookups["budget"] = build_lookup_sheet(
        wb, "Budgets", data.get("budgets") or [],
        [
            ("UUID", lambda b: b.get("uuid")),
            ("Budget Name", lambda b: b.get("name")),
            ("Amount", lambda b: b.get("amount")),
            ("Period", lambda b: b.get("period")),
        ],
    )
    lookups["loan"] = build_lookup_sheet(
        wb, "Loans", data.get("loans") or [],
        [
            ("UUID", lambda l: l.get("uuid")),
            ("Loan Name", lambda l: l.get("name")),
            ("Amount", lambda l: l.get("amount")),
        ],
    )

    # how many tag slots does this backup actually need?
    max_tags = max((len(t.get("tags") or []) for t in transactions), default=0)

    # ---------------------------------------------------------------
    # Transactions sheet
    # ---------------------------------------------------------------
    ws = wb.create_sheet("Transactions", 0)

    # (visible header, txn-field, lookup key) for single-value entity refs
    entity_cols = [
        ("Account", "account", "account"),
        ("Category", "category", "category"),
        ("From Account", "fromAccount", "account"),
        ("To Account", "toAccount", "account"),
        ("Place", "place", "place"),
        ("Person", "people", "people"),
        ("Recurring", "recurring", "recurring"),
        ("Budget", "budget", "budget"),
        ("Loan", "loan", "loan"),
        ("User", "user", "user"),
    ]
    # only keep entity columns whose lookup sheet actually exists
    entity_cols = [ec for ec in entity_cols if lookups.get(ec[2])]

    tag_cols = [f"Tag {i+1}" for i in range(max_tags)] if lookups.get("tag") else []

    visible_headers = (
        ["ID", "Name", "Amount", "Type", "Created At", "Updated At", "Description"]
        + [h for h, _, _ in entity_cols]
        + tag_cols
    )
    n_visible = len(visible_headers)

    helper_headers = [f"{h} UUID" for h, _, _ in entity_cols] + [
        f"Tag {i+1} UUID" for i in range(max_tags)
    ]
    all_headers = visible_headers + helper_headers
    ws.append(all_headers)
    style_header(ws, len(all_headers))

    helper_start = n_visible + 1
    helper_col_of = {
        name: helper_start + i for i, name in enumerate(
            [h for h, _, _ in entity_cols] + tag_cols
        )
    }

    def col_letter(idx):
        return get_column_letter(idx)

    def lookup_formula(helper_col, sheet_name, n_rows):
        hc = col_letter(helper_col)
        return (
            f'=IF({hc}{{r}}="","",'
            f'IFERROR(INDEX({sheet_name}!$B$2:$B${n_rows},'
            f'MATCH({hc}{{r}},{sheet_name}!$A$2:$A${n_rows},0)),"Unknown"))'
        )

    row = 2
    for t in transactions:
        r = row
        ws.cell(row=r, column=1, value=t.get("id"))
        ws.cell(row=r, column=2, value=t.get("name"))
        ws.cell(row=r, column=3, value=t.get("amount"))
        ws.cell(row=r, column=4, value=txn_type_map.get(t.get("type"), t.get("type")))
        ws.cell(row=r, column=5, value=t.get("createdAt"))
        ws.cell(row=r, column=6, value=t.get("updatedAt"))
        ws.cell(row=r, column=7, value=t.get("description") or "")

        col = 8
        for header, field, lookup_key in entity_cols:
            hc = helper_col_of[header]
            ws.cell(row=r, column=hc, value=t.get(field) or "")
            sheet_name, n_rows = lookups[lookup_key]
            ws.cell(row=r, column=col, value=lookup_formula(hc, sheet_name, n_rows).format(r=r))
            col += 1

        tags_list = t.get("tags") or []
        for i, tag_col_name in enumerate(tag_cols):
            hc = helper_col_of[tag_col_name]
            uuid_val = tags_list[i] if i < len(tags_list) else ""
            ws.cell(row=r, column=hc, value=uuid_val)
            sheet_name, n_rows = lookups["tag"]
            ws.cell(row=r, column=col, value=lookup_formula(hc, sheet_name, n_rows).format(r=r))
            col += 1

        row += 1

    n_tx = len(transactions)
    apply_font_and_border(ws, n_tx, len(all_headers))
    for r in range(2, 2 + n_tx):
        ws.cell(row=r, column=3).number_format = "#,##0.00"

    autosize(ws, [7, 26, 12, 10, 20, 20, 24] + [20] * (n_visible - 7) + [38] * len(helper_headers))
    ws.freeze_panes = "C2"

    for c in range(helper_start, helper_start + len(helper_headers)):
        ws.column_dimensions[get_column_letter(c)].hidden = True

    ws.auto_filter.ref = f"A1:{get_column_letter(n_visible)}{1 + n_tx}"

    wb.save(output_path)
    print(f"Saved {output_path} — {n_tx} transactions, "
          f"{sum(1 for v in lookups.values() if v)} lookup sheets.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    in_path = Path(sys.argv[1])
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else in_path.with_suffix(".xlsx")
    convert(in_path, out_path)
